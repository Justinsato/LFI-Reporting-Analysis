import os
import pandas as pd
import requests
import json
from datetime import datetime
from dotenv import load_dotenv
import logging
from openpyxl import load_workbook, Workbook
import openai

# Load environment variables
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
TENANT_ID = os.getenv("TENANT_ID")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
SHAREPOINT_SITE_ID = os.getenv("SHAREPOINT_SITE_ID")
SHAREPOINT_TASK_LIST_ID = os.getenv("SHAREPOINT_TASK_LIST_ID")

openai.api_key = OPENAI_API_KEY

# Setup logging
os.makedirs("logs", exist_ok=True)
logfile = f"logs/variance_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    filename=logfile,
    level=logging.INFO,
    format="%(asctime)s — %(levelname)s — %(message)s"
)

def log(msg):
    print(msg)
    logging.info(msg)

def get_graph_token():
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    data = {
        "client_id": CLIENT_ID,
        "scope": "https://graph.microsoft.com/.default",
        "client_secret": CLIENT_SECRET,
        "grant_type": "client_credentials"
    }
    response = requests.post(url, headers=headers, data=data)
    response.raise_for_status()
    return response.json()["access_token"]

def post_task_to_sharepoint(title, body, token):
    url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_ID}/lists/{SHAREPOINT_TASK_LIST_ID}/items"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    data = {
        "fields": {
            "Title": title,
            "Description": body,
            "Status": "Not Started"
        }
    }
    try:
        response = requests.post(url, headers=headers, json=data)
        if response.status_code == 201:
            log(f"✅ Posted task: {title}")
        else:
            log(f"⚠️ Failed to post task '{title}': {response.status_code} - {response.text}")
    except Exception as e:
        logging.error(f"Error posting to SharePoint: {e}")

def load_supporting_data(directory):
    combined_data = []
    for file in os.listdir(directory):
        if file.endswith(".csv"):
            file_path = os.path.join(directory, file)
            try:
                df = pd.read_csv(file_path, dtype=str, encoding='utf-8', on_bad_lines='skip', engine='python')
                df['source_file'] = file
                combined_data.append(df)
            except UnicodeDecodeError:
                try:
                    df = pd.read_csv(file_path, dtype=str, encoding='ISO-8859-1', on_bad_lines='skip', engine='python')
                    df['source_file'] = file
                    combined_data.append(df)
                except Exception as e:
                    logging.warning(f"Skipped file due to encoding issue: {file} - {e}")
            except Exception as e:
                logging.warning(f"Skipped file due to error: {file} - {e}")
    return pd.concat(combined_data, ignore_index=True) if combined_data else pd.DataFrame()

def generate_commentary(account, old_comment, support_data):
    account = str(account).strip().replace('"', "'")
    old_comment = str(old_comment or "").strip().replace('"', "'")

    logging.info(f"🔍 Generating note for account: {account} | Existing comment: {old_comment}")

    relevant_data = support_data[support_data.apply(
        lambda row: str(account).lower() in str(row.values).lower(), axis=1)]

    support_summary = relevant_data.head(5).to_string(index=False) if not relevant_data.empty else "No matching data."

    prompt = f'''
You are a seasoned financial analyst generating root-cause-level actual-to-budget variance commentary for a multifamily apartment portfolio. Each comment must be insightful, conclusive, concise, and grounded in available support data. Do not suggest further analysis and do not say things that make you look uninformed.  Speak in the voice of a confident operator preparing content for investor and executive consumption. Avoid generic language and repetition.  Speak conversationaly.  I want you to go back through each comment you make at least once and revise it to be even more concise and punchy.

Account: {account}
Original comment: "{old_comment}"
Relevant Support Data (sample rows):
{support_summary}

Write the explanation that will be embedded into the Excel report.
'''

    try:
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": prompt}]
        )
        content = response.choices[0].message.content.strip()
        return content or "⚠️ GPT returned no response"
    except Exception as e:
        logging.error(f"OpenAI error for account {account}: {e}")
        return "⚠️ GPT ERROR"

def generate_analysis_outputs(df):
    combined_context = "\n".join(
        row.get('Variance Explanation', '') for _, row in df.iterrows() if pd.notna(row.get('Variance Explanation')))
    sections = [
        ("Top 10 Task List", "Create a Top 10 Task List for property managers based on these variance explanations."),
        ("Top 10 Risk List", "From this same data, list the Top 10 Risks to future financial performance."),
        ("Top 10 Opportunity Areas", "Identify the Top 10 areas of opportunity based on this data for financial or operational improvement."),
        ("Investor Concerns", "From all the items above, highlight anything that would be concerning from an investor's perspective."),
        ("Executive Summary", "Write a short, clear, executive summary of the financial results this period. Favor the property operator's perspective and frame negatives as areas of growth or focus.")
    ]

    result = {}
    for title, prompt in sections:
        full_prompt = f"{prompt}\n\n{combined_context}"
        try:
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": full_prompt}]
            )
            result[title] = response.choices[0].message.content.strip()
        except Exception as e:
            result[title] = f"⚠️ ERROR: {e}"
            logging.error(f"Error generating section '{title}': {e}")
    return result

def main():
    INPUT_FILE = r"C:\\Users\\JustinSato\\OneDrive - Left Field Investments\\Documents - LFI Data and Reporting\\Monthly Variance Files\\BAPF-A Consolidate Financials.xlsx"
    OUTPUT_FILE = r"C:\\Users\\JustinSato\\OneDrive - Left Field Investments\\Documents - LFI Data and Reporting\\AI Outputs\\BAPF-A Consolidate Financials w_notes.xlsx"
    OUTPUT_SUMMARY_FILE = OUTPUT_FILE.replace(".xlsx", "_summary.xlsx")
    SUPPORTING_DATA_DIR = r"C:\\Users\\JustinSato\\OneDrive - Left Field Investments\\Documents - LFI Data and Reporting\\Monthly Variance Files\\Supporting Files"

    try:
        wb = load_workbook(INPUT_FILE)
        ws = wb["Con w Narrative"]
        log("✅ Loaded Excel file")
    except Exception as e:
        logging.error("❌ Failed to load Excel file: {}".format(e))
        return

    support_data = load_supporting_data(SUPPORTING_DATA_DIR)
    updated_count = 0

    for row in range(11, ws.max_row + 1):
        variance = ws.cell(row=row, column=9).value  # Column I
        account = ws.cell(row=row, column=3).value   # Column C
        comment = ws.cell(row=row, column=11).value  # Column K

        if variance not in [None, ""]:
            log(f"➡️ Processing row {row} | Account: {account}")
            if not comment or comment.strip() == "":
                new_comment = generate_commentary(account, comment, support_data)
                if new_comment:
                    ws.cell(row=row, column=11).value = new_comment
                    updated_count += 1
                    log(f"📝 Added comment to row {row}")
                else:
                    log(f"⚠️ No comment generated for row {row}")
            else:
                log(f"⏭️ Skipping row {row} — comment already exists")

    try:
        wb.save(OUTPUT_FILE)
        log("✅ Saved updated file to: {} with {} new variance notes".format(OUTPUT_FILE, updated_count))
    except Exception as e:
        logging.error("❌ Failed to save Excel file: {}".format(e))
        return

    token = get_graph_token()
    df = pd.read_excel(OUTPUT_FILE, sheet_name="Con w Narrative", header=6)
    summary_sections = generate_analysis_outputs(df)

    wb_summary = Workbook()
    for title, content in summary_sections.items():
        ws_section = wb_summary.create_sheet(title=title[:31])
        for i, line in enumerate(content.splitlines(), start=1):
            ws_section.cell(row=i, column=1).value = line
    if 'Sheet' in wb_summary.sheetnames:
        del wb_summary['Sheet']
    wb_summary.save(OUTPUT_SUMMARY_FILE)
    log(f"✅ Summary file saved to: {OUTPUT_SUMMARY_FILE}")

if __name__ == "__main__":
    main()
